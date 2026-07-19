# -*- coding: utf-8 -*-
"""
build_planilha_notinha.py — gera a Notinha_Planilha_Mestre.xlsx

Planilha financeira anual da Notinha: 1 arquivo com 12 abas mensais
(Janeiro..Dezembro) + Como usar + Visão Anual + Metas Financeiras + Investimento.

Categorias: fonte de verdade é o worker `processar-fila` (edge function),
const CATEGORIAS — 12 categorias. Forma de pagamento vem como texto livre da
nota, então é normalizada aqui em 6 grupos.

Uso:
    python3 build_planilha_notinha.py [saida.xlsx] [--dados dados.json] [--ano AAAA]

`--dados` aponta para um JSON com as notas reais (fora do repositório, para não
versionar dados pessoais). Sem ele, a planilha sai como TEMPLATE com um punhado
de lançamentos genéricos de exemplo.
"""
import argparse
import json
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- identidade
TEAL = "288A89"      # cabeçalhos/títulos
DARK = "0D1117"      # faixas de destaque
OFFWHITE = "F4F6F6"  # células/cartões
BORDER = "E3E9E9"    # bordas
WHITE = "FFFFFF"

F_TEAL = PatternFill("solid", fgColor=TEAL)
F_DARK = PatternFill("solid", fgColor=DARK)
F_CARD = PatternFill("solid", fgColor=OFFWHITE)

FT_TITLE = Font(name="Arial", size=18, bold=True, color=WHITE)
FT_HEAD = Font(name="Arial", size=11, bold=True, color=WHITE)
FT_BODY = Font(name="Arial", size=10, color="1A1A1A")
FT_BODY_B = Font(name="Arial", size=10, bold=True, color="1A1A1A")
FT_BIG = Font(name="Arial", size=16, bold=True, color=WHITE)

THIN = Side(style="thin", color=BORDER)
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)
AR = Alignment(horizontal="right", vertical="center")

FMT_BRL = 'R$ #,##0.00'
FMT_DATE = 'DD/MM/YYYY'
FMT_PCT = '0.0%'

MESES = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
MESES_ABREV = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
               "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

# Fonte de verdade: const CATEGORIAS do worker processar-fila
CATEGORIAS = ["Alimentação", "Mercado", "Farmácia", "Combustível", "Transporte",
              "Vestuário", "Eletrônicos", "Casa", "Saúde", "Lazer", "Serviços", "Outros"]
# Grupos de forma de pagamento (a nota traz texto livre; normalizamos nestes)
FORMAS = ["Débito", "Crédito", "Pix", "Dinheiro", "Voucher", "Outros"]
SAIDAS_LINHAS = FORMAS + ["Reserva", "Metas"]

# Cor de cada "pill" (fundo, texto) — categorias
CAT_COR = {
    "Mercado": ("1F7A6D", WHITE), "Alimentação": ("E8743B", WHITE),
    "Farmácia": ("2E8B57", WHITE), "Combustível": ("B7791F", WHITE),
    "Transporte": ("2B6CB0", WHITE), "Vestuário": ("805AD5", WHITE),
    "Eletrônicos": ("4C51BF", WHITE), "Casa": ("2C7A7B", WHITE),
    "Saúde": ("C53030", WHITE), "Lazer": ("D53F8C", WHITE),
    "Serviços": ("4A5568", WHITE), "Outros": ("A0AEC0", "1A202C"),
}
# Cor de cada "pill" — formas de pagamento
PAG_COR = {
    "Débito": ("3182CE", WHITE), "Crédito": ("805AD5", WHITE),
    "Pix": ("319795", WHITE), "Dinheiro": ("38A169", WHITE),
    "Voucher": ("DD6B20", WHITE), "Outros": ("A0AEC0", "1A202C"),
}

DV_CATS = '"' + ",".join(CATEGORIAS) + '"'
DV_FORMAS = '"' + ",".join(FORMAS) + '"'
DV_PAGO = '"Sim,Não"'

GASTOS_INI, GASTOS_FIM = 5, 64          # 60 linhas de lançamentos (col A:E)
CAT_INI = 5
CAT_FIM = CAT_INI + len(CATEGORIAS) - 1  # 16


# ---------------------------------------------------------------- normalização
def _sem_acento(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn").lower().strip()


def norm_categoria(c):
    if not c:
        return "Outros"
    k = _sem_acento(c)
    mapa = {
        "alimentacao": "Alimentação", "padaria": "Alimentação",
        "mercado": "Mercado", "farmacia": "Farmácia", "combustivel": "Combustível",
        "transporte": "Transporte", "aluguel carro": "Transporte",
        "vestuario": "Vestuário", "eletronicos": "Eletrônicos", "casa": "Casa",
        "saude": "Saúde", "lazer": "Lazer", "servicos": "Serviços", "outros": "Outros",
    }
    return mapa.get(k, "Outros")


def norm_forma(f):
    if not f:
        return "Outros"
    s = _sem_acento(f)
    if "credito" in s or "credit" in s or "nubank" in s:
        return "Crédito"
    if "ifood" in s or "voucher" in s or "alelo" in s or "stix" in s or "vale" in s:
        return "Voucher"
    if "debito" in s or "debit" in s:
        return "Débito"
    if "pix" in s:
        return "Pix"
    if "dinheiro" in s:
        return "Dinheiro"
    if "mastercard" in s or "cartao" in s or "pos" in s or "tef" in s:
        return "Débito"
    return "Outros"


# ---------------------------------------------------------------- helpers
def paint(ws, rng, fill=None, font=None, border=B_ALL, align=None, fmt=None):
    for row in ws[rng]:
        for c in row:
            if fill is not None:
                c.fill = fill
            if font is not None:
                c.font = font
            if border is not None:
                c.border = border
            if align is not None:
                c.alignment = align
            if fmt is not None:
                c.number_format = fmt


def merged(ws, rng, text, fill=F_TEAL, font=FT_HEAD, align=AC, fmt=None):
    ws.merge_cells(rng)
    ws[rng.split(":")[0]] = text
    paint(ws, rng, fill=fill, font=font, align=align, fmt=fmt)


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def card_rows(ws, rng, fmt_map=None):
    paint(ws, rng, fill=F_CARD, font=FT_BODY, align=AL)
    if fmt_map:
        first, last = rng.split(":")
        r1 = int("".join(ch for ch in first if ch.isdigit()))
        r2 = int("".join(ch for ch in last if ch.isdigit()))
        for col, fmt in fmt_map.items():
            for r in range(r1, r2 + 1):
                ws[f"{col}{r}"].number_format = fmt
                if fmt in (FMT_BRL, FMT_PCT):
                    ws[f"{col}{r}"].alignment = AR
                elif fmt == FMT_DATE:
                    ws[f"{col}{r}"].alignment = AC


def pill(cell, valor, cor_map):
    """Pinta uma célula como 'pill' colorido conforme o valor."""
    bg, fg = cor_map.get(valor, ("A0AEC0", "1A202C"))
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=10, bold=True, color=fg)
    cell.alignment = AC
    cell.border = B_ALL


def cf_pills(ws, rng, cor_map):
    """Formatação condicional: colore automaticamente cada valor no range
    (assim, lançamentos futuros da Notinha já saem com o pill certo)."""
    for valor, (bg, fg) in cor_map.items():
        dxf = DifferentialStyle(
            fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
            font=Font(color=fg, bold=True))
        rule = Rule(type="containsText", operator="containsText", text=valor, dxf=dxf)
        # fórmula exigida pelo Excel para containsText
        first = rng.split(":")[0]
        rule.formula = [f'NOT(ISERROR(SEARCH("{valor}",{first})))']
        ws.conditional_formatting.add(rng, rule)


# ---------------------------------------------------------------- aba mensal
def montar_mes(ws, idx, dados=None):
    mes = MESES[idx]
    ano = dados["ano"] if dados else datetime.now().year
    ws.sheet_properties.tabColor = TEAL
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 30, "B": 12, "C": 12, "D": 15, "E": 14, "F": 2,
                    "G": 13, "H": 15, "I": 13, "J": 2,
                    "K": 17, "L": 13, "M": 12, "N": 12, "O": 2})
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A5"

    merged(ws, "A1:N1", f"{mes} · Notinha {ano}", fill=F_TEAL, font=FT_TITLE)

    # ---- Gastos do Mês (A:E) — lançados pela Notinha
    merged(ws, "A3:C3", "GASTOS DO MÊS — lançados pela Notinha")
    ws["D3"] = "Total do mês"
    ws["E3"] = f"=SUM(E{GASTOS_INI}:E{GASTOS_FIM})"
    paint(ws, "D3:E3", fill=F_DARK, font=FT_BIG, align=AR)
    ws["D3"].font = FT_HEAD
    ws["D3"].alignment = AC
    ws["E3"].number_format = FMT_BRL
    for col, h in zip("ABCDE", ["Nome", "Data", "Forma", "Categoria", "Valor"]):
        ws[f"{col}4"] = h
    paint(ws, "A4:E4", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, f"A{GASTOS_INI}:E{GASTOS_FIM}", {"B": FMT_DATE, "E": FMT_BRL})
    # pills automáticos (Forma = col C, Categoria = col D)
    cf_pills(ws, f"C{GASTOS_INI}:C{GASTOS_FIM}", PAG_COR)
    cf_pills(ws, f"D{GASTOS_INI}:D{GASTOS_FIM}", CAT_COR)

    # preenche com dados reais (ou exemplo)
    linhas = _linhas_do_mes(dados, idx)
    for i, (nome, data, forma, cat, valor) in enumerate(linhas[:GASTOS_FIM - GASTOS_INI + 1]):
        r = GASTOS_INI + i
        ws[f"A{r}"] = nome
        ws[f"B{r}"] = data
        ws[f"B{r}"].number_format = FMT_DATE
        ws[f"B{r}"].alignment = AC
        ws[f"C{r}"] = forma
        pill(ws[f"C{r}"], forma, PAG_COR)
        ws[f"D{r}"] = cat
        pill(ws[f"D{r}"], cat, CAT_COR)
        ws[f"E{r}"] = valor
        ws[f"E{r}"].number_format = FMT_BRL
        ws[f"E{r}"].alignment = AR

    # ---- Fixos (G:I) — você preenche
    merged(ws, "G3:H3", "FIXOS (você preenche)")
    ws["I3"] = "=SUM(I5:I12)"
    paint(ws, "I3:I3", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    for col, h in zip("GHI", ["Forma", "Categoria", "Valor"]):
        ws[f"{col}4"] = h
    paint(ws, "G4:I4", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, "G5:I12", {"I": FMT_BRL})

    # ---- Cartão de Crédito (G:I)
    merged(ws, "G14:H14", "CARTÃO DE CRÉDITO (você preenche)")
    ws["I14"] = "=SUM(I16:I23)"
    paint(ws, "I14:I14", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    for col, h in zip("GHI", ["Forma", "Categoria", "Valor"]):
        ws[f"{col}15"] = h
    paint(ws, "G15:I15", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, "G16:I23", {"I": FMT_BRL})

    # ---- Entradas
    merged(ws, "G25:I25", "ENTRADAS (você preenche)")
    for r, label in ((26, "Salário"), (27, "Bônus"), (28, "Total")):
        merged(ws, f"G{r}:H{r}", label, fill=F_CARD, font=FT_BODY_B, align=AL)
    card_rows(ws, "I26:I27", {"I": FMT_BRL})
    ws["I28"] = "=SUM(I26:I27)"
    paint(ws, "G28:I28", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    ws["G28"].alignment = AL
    ent = (dados or {}).get("entradas", {}).get(str(idx + 1)) if dados else None
    if ent:
        ws["I26"] = ent.get("salario", 0) or 0
        ws["I27"] = ent.get("bonus", 0) or 0

    # ---- Saídas por forma de pagamento (já saiu no mês) → alimenta o Saldo
    merged(ws, "G30:I30", "SAÍDAS POR FORMA DE PAGAMENTO")
    for i, label in enumerate(SAIDAS_LINHAS):
        r = 31 + i
        ws.merge_cells(f"G{r}:H{r}")
        ws[f"G{r}"] = label
        if label in PAG_COR:
            bg, fg = PAG_COR[label]
            paint(ws, f"G{r}:H{r}", fill=PatternFill("solid", fgColor=bg),
                  font=Font(name="Arial", size=10, bold=True, color=fg), align=AC)
            ws[f"I{r}"] = (
                f"=SUMIF($C${GASTOS_INI}:$C${GASTOS_FIM},$G{r},$E${GASTOS_INI}:$E${GASTOS_FIM})"
                f"+SUMIF($G$5:$G$12,$G{r},$I$5:$I$12)"
                f"+SUMIF($G$16:$G$23,$G{r},$I$16:$I$23)"
            )
        else:  # Reserva e Metas: quanto você guardou (manual)
            paint(ws, f"G{r}:H{r}", fill=F_CARD, font=FT_BODY, align=AL)
        paint(ws, f"I{r}:I{r}", fill=F_CARD, font=FT_BODY, align=AR, fmt=FMT_BRL)
    tot_r = 31 + len(SAIDAS_LINHAS)
    merged(ws, f"G{tot_r}:H{tot_r}", "Total de saídas", fill=F_DARK, font=FT_HEAD, align=AL)
    ws[f"I{tot_r}"] = f"=SUM(I31:I{tot_r - 1})"
    paint(ws, f"I{tot_r}:I{tot_r}", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)

    # ---- Investimentos
    inv_t = tot_r + 2
    merged(ws, f"G{inv_t}:I{inv_t}", "INVESTIMENTOS")
    for k, label in enumerate(("Reserva", "Renda fixa")):
        r = inv_t + 1 + k
        merged(ws, f"G{r}:H{r}", label, fill=F_CARD, font=FT_BODY, align=AL)
    card_rows(ws, f"I{inv_t + 1}:I{inv_t + 2}", {"I": FMT_BRL})
    inv_tot = inv_t + 3
    merged(ws, f"G{inv_tot}:H{inv_tot}", "Total", fill=F_DARK, font=FT_HEAD, align=AL)
    ws[f"I{inv_tot}"] = f"=SUM(I{inv_t + 1}:I{inv_t + 2})"
    paint(ws, f"I{inv_tot}:I{inv_tot}", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)

    # ---- Saldo (destaque)
    sal_t = inv_tot + 2
    merged(ws, f"G{sal_t}:I{sal_t}", "SALDO DO MÊS  (Entradas − Saídas)")
    ws.row_dimensions[sal_t + 1].height = 32
    merged(ws, f"G{sal_t + 1}:I{sal_t + 1}", f"=I28-I{tot_r}",
           fill=F_DARK, font=FT_BIG, align=AC, fmt=FMT_BRL)

    # ---- Gastos do mês por categoria (K:N)
    merged(ws, "K3:N3", "GASTOS DO MÊS POR CATEGORIA")
    for col, h in zip("KLMN", ["Categoria", "Valor esperado", "Valor gasto", "%"]):
        ws[f"{col}4"] = h
    paint(ws, "K4:N4", fill=F_TEAL, font=FT_HEAD, align=AC)
    for i, cat in enumerate(CATEGORIAS):
        r = CAT_INI + i
        ws[f"K{r}"] = cat
        pill(ws[f"K{r}"], cat, CAT_COR)
        ws[f"M{r}"] = (f"=SUMIF($D${GASTOS_INI}:$D${GASTOS_FIM},$K{r},"
                       f"$E${GASTOS_INI}:$E${GASTOS_FIM})")
        ws[f"N{r}"] = f"=IFERROR(M{r}/$E$3,0)"
    card_rows(ws, f"L{CAT_INI}:N{CAT_FIM}", {"L": FMT_BRL, "M": FMT_BRL, "N": FMT_PCT})
    for i in range(len(CATEGORIAS)):  # restaura borda das células pill em K
        ws[f"K{CAT_INI + i}"].border = B_ALL
    r = CAT_FIM + 1
    ws[f"K{r}"] = "Total"
    ws[f"L{r}"] = f"=SUM(L{CAT_INI}:L{CAT_FIM})"
    ws[f"M{r}"] = f"=SUM(M{CAT_INI}:M{CAT_FIM})"
    ws[f"N{r}"] = f"=IFERROR(M{r}/$E$3,0)"
    paint(ws, f"K{r}:N{r}", fill=F_DARK, font=FT_HEAD, align=AR)
    ws[f"K{r}"].alignment = AL
    ws[f"L{r}"].number_format = FMT_BRL
    ws[f"M{r}"].number_format = FMT_BRL
    ws[f"N{r}"].number_format = FMT_PCT

    # ---- Contas a pagar (faturas) — checklist do que ainda falta pagar
    merged(ws, "K19:N19", "CONTAS A PAGAR (faturas)")
    for col, h in zip("KLMN", ["Descrição", "Valor", "Vencimento", "Pago?"]):
        ws[f"{col}20"] = h
    paint(ws, "K20:N20", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, "K21:N30", {"L": FMT_BRL, "M": FMT_DATE})

    # ---- Insights do mês (K:N)
    merged(ws, "K32:N32", "INSIGHTS DO MÊS")
    g1, g2 = GASTOS_INI, GASTOS_FIM
    insights = [
        ("Nº de lançamentos", f"=COUNT($E${g1}:$E${g2})", None),
        ("Maior compra", f"=IF(COUNT($E${g1}:$E${g2})=0,0,MAX($E${g1}:$E${g2}))", FMT_BRL),
        ("Onde foi a maior compra",
         f'=IF($E$3=0,"—",INDEX($A${g1}:$A${g2},MATCH(MAX($E${g1}:$E${g2}),$E${g1}:$E${g2},0)))', None),
        ("Categoria campeã",
         f'=IF($E$3=0,"—",INDEX($K${CAT_INI}:$K${CAT_FIM},MATCH(MAX($M${CAT_INI}:$M${CAT_FIM}),$M${CAT_INI}:$M${CAT_FIM},0)))', None),
        ("Ticket médio", f"=IFERROR(AVERAGE($E${g1}:$E${g2}),0)", FMT_BRL),
        ("Gasto médio por dia",
         f"=IFERROR($E$3/DAY(EOMONTH(DATE({ano},{idx + 1},1),0)),0)", FMT_BRL),
        ("% das entradas já gasto", f"=IFERROR($I${tot_r}/$I$28,0)", FMT_PCT),
    ]
    for i, (label, formula, fmt) in enumerate(insights):
        r = 33 + i
        merged(ws, f"K{r}:M{r}", label, fill=F_CARD, font=FT_BODY_B, align=AL)
        ws[f"N{r}"] = formula
        paint(ws, f"N{r}:N{r}", fill=F_CARD, font=FT_BODY, align=AR, fmt=fmt or "General")

    # ---- Gráfico de pizza
    pie = PieChart()
    pie.title = "Gastos por categoria"
    data = Reference(ws, min_col=13, min_row=4, max_row=CAT_FIM)
    cats = Reference(ws, min_col=11, min_row=CAT_INI, max_row=CAT_FIM)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.height, pie.width = 8.5, 12
    ws.add_chart(pie, "P32")

    # ---- validações (dropdowns)
    dv_cat = DataValidation(type="list", formula1=DV_CATS, allow_blank=True)
    dv_cat.add(f"D{GASTOS_INI}:D{GASTOS_FIM}")
    dv_cat.add("H5:H12")
    dv_cat.add("H16:H23")
    dv_forma = DataValidation(type="list", formula1=DV_FORMAS, allow_blank=True)
    dv_forma.add(f"C{GASTOS_INI}:C{GASTOS_FIM}")
    dv_forma.add("G5:G12")
    dv_forma.add("G16:G23")
    dv_pago = DataValidation(type="list", formula1=DV_PAGO, allow_blank=True)
    dv_pago.add("N21:N30")
    for dv in (dv_cat, dv_forma, dv_pago):
        ws.add_data_validation(dv)


def _linhas_do_mes(dados, idx):
    """Retorna [(nome, date, forma, categoria, valor)] do mês idx (0..11)."""
    if not dados:
        if idx == 0:  # exemplo genérico só em Janeiro (template)
            ano = datetime.now().year
            return [
                ("Supermercado (exemplo)", datetime(ano, 1, 5), "Débito", "Mercado", 312.45),
                ("Restaurante (exemplo)", datetime(ano, 1, 8), "Crédito", "Alimentação", 56.90),
                ("Transporte (exemplo)", datetime(ano, 1, 12), "Pix", "Transporte", 18.75),
                ("Conta de casa (exemplo)", datetime(ano, 1, 15), "Débito", "Casa", 187.30),
                ("Farmácia (exemplo)", datetime(ano, 1, 20), "Dinheiro", "Farmácia", 74.20),
            ]
        return []
    out = []
    for d, estab, forma, cat, valor in dados["notas"]:
        dt = datetime.strptime(d, "%Y-%m-%d")
        if dt.month != idx + 1:
            continue
        out.append((estab or "—", dt, norm_forma(forma), norm_categoria(cat), float(valor)))
    return out


# ---------------------------------------------------------------- visão anual
def montar_visao_anual(ws, dados=None):
    ws.sheet_properties.tabColor = DARK
    ws.sheet_view.showGridLines = False
    ano = dados["ano"] if dados else datetime.now().year
    set_widths(ws, {"A": 22, **{get_column_letter(c): 12 for c in range(2, 15)}})
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A5"
    merged(ws, "A1:N1", f"Visão Anual · Notinha {ano}", fill=F_TEAL, font=FT_TITLE)

    merged(ws, "A3:D3", "RESUMO MENSAL")
    for col, h in zip("ABCD", ["Mês", "Total Entradas", "Total Saídas", "Saldo"]):
        ws[f"{col}4"] = h
    paint(ws, "A4:D4", fill=F_TEAL, font=FT_HEAD, align=AC)
    # descobre a linha do "Total de saídas" e do "saldo" na aba mensal
    n_saidas = len(SAIDAS_LINHAS)
    lin_tot_saidas = 31 + n_saidas          # I{lin} = total saídas
    lin_saldo = lin_tot_saidas + 2 + 3 + 2 + 1  # segue o empilhamento de montar_mes
    for i, mes in enumerate(MESES):
        r = 5 + i
        ws[f"A{r}"] = mes
        ws[f"B{r}"] = f"={mes}!$I$28"
        ws[f"C{r}"] = f"={mes}!$I${lin_tot_saidas}"
        ws[f"D{r}"] = f"={mes}!$I$28-{mes}!$I${lin_tot_saidas}"
    card_rows(ws, "A5:D16", {"B": FMT_BRL, "C": FMT_BRL, "D": FMT_BRL})
    ws["A17"] = "TOTAL DO ANO"
    for col in "BCD":
        ws[f"{col}17"] = f"=SUM({col}5:{col}16)"
    paint(ws, "A17:D17", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    ws["A17"].alignment = AL
    ws["A17"].number_format = "General"

    bar = BarChart()
    bar.type = "col"
    bar.title = "Gasto total por mês"
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=16), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=5, max_row=16))
    bar.legend = None
    bar.height, bar.width = 7.5, 15
    ws.add_chart(bar, "F3")

    # Categoria × 12 meses
    merged(ws, "A20:N20", "GASTOS POR CATEGORIA × MÊS")
    ws["A21"] = "Categoria"
    for i, ab in enumerate(MESES_ABREV):
        ws[f"{get_column_letter(2 + i)}21"] = ab
    ws["N21"] = "Total"
    paint(ws, "A21:N21", fill=F_TEAL, font=FT_HEAD, align=AC)
    for i, cat in enumerate(CATEGORIAS):
        r = 22 + i
        ws[f"A{r}"] = cat
        pill(ws[f"A{r}"], cat, CAT_COR)
        for m, mes in enumerate(MESES):
            ws[f"{get_column_letter(2 + m)}{r}"] = f"={mes}!$M${CAT_INI + i}"
        ws[f"N{r}"] = f"=SUM(B{r}:M{r})"
    card_rows(ws, "B22:N33", {get_column_letter(c): FMT_BRL for c in range(2, 15)})
    for i in range(len(CATEGORIAS)):
        ws[f"A{22 + i}"].border = B_ALL
    ws["A34"] = "Total"
    for c in range(2, 15):
        col = get_column_letter(c)
        ws[f"{col}34"] = f"=SUM({col}22:{col}33)"
    paint(ws, "A34:N34", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    ws["A34"].alignment = AL
    ws["A34"].number_format = "General"

    # Top produtos / maiores aumentos (diferencial item-a-item da Notinha)
    merged(ws, "A37:C37", "TOP PRODUTOS MAIS COMPRADOS")
    for col, h in zip("ABC", ["Produto", "Vezes", "Gasto total"]):
        ws[f"{col}38"] = h
    paint(ws, "A38:C38", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, "A39:C48", {"C": FMT_BRL})
    for i, (prod, vezes, gasto) in enumerate((dados or {}).get("top_produtos", [])[:10]):
        r = 39 + i
        ws[f"A{r}"] = prod
        ws[f"B{r}"] = vezes
        ws[f"B{r}"].alignment = AC
        ws[f"C{r}"] = gasto

    merged(ws, "E37:H37", "MAIORES VARIAÇÕES DE PREÇO")
    for col, h in zip("EFGH", ["Produto", "Antes", "Agora", "Variação"]):
        ws[f"{col}38"] = h
    paint(ws, "E38:H38", fill=F_TEAL, font=FT_HEAD, align=AC)
    card_rows(ws, "E39:H48", {"F": FMT_BRL, "G": FMT_BRL})
    for i, (prod, antes, agora, di, df, pct) in enumerate((dados or {}).get("aumentos", [])[:10]):
        r = 39 + i
        ws[f"E{r}"] = prod
        ws[f"F{r}"] = antes
        ws[f"G{r}"] = agora
        ws[f"H{r}"] = pct / 100.0
        ws[f"H{r}"].number_format = FMT_PCT
        ws[f"H{r}"].alignment = AR
        ws[f"H{r}"].font = Font(name="Arial", size=10, bold=True,
                                color="C53030" if pct > 0 else "2E8B57")

    # Insights do ano
    merged(ws, "A51:D51", "INSIGHTS DO ANO")
    insights = [
        ("Mês com maior gasto",
         '=IF(MAX(C5:C16)=0,"—",INDEX(A5:A16,MATCH(MAX(C5:C16),C5:C16,0)))', None),
        ("Mês com maior saldo",
         '=IF(SUM(B5:B16)=0,"—",INDEX(A5:A16,MATCH(MAX(D5:D16),D5:D16,0)))', None),
        ("Categoria campeã do ano",
         '=IF(SUM(N22:N33)=0,"—",INDEX(A22:A33,MATCH(MAX(N22:N33),N22:N33,0)))', None),
        ("Média mensal de saídas", '=IFERROR(AVERAGEIF(C5:C16,">0"),0)', FMT_BRL),
        ("Total investido no ano", "=Investimento!$D$17", FMT_BRL),
        ("Taxa de poupança (Saldo ÷ Entradas)", "=IFERROR(D17/B17,0)", FMT_PCT),
    ]
    for i, (label, formula, fmt) in enumerate(insights):
        r = 52 + i
        merged(ws, f"A{r}:C{r}", label, fill=F_CARD, font=FT_BODY_B, align=AL)
        ws[f"D{r}"] = formula
        paint(ws, f"D{r}:D{r}", fill=F_CARD, font=FT_BODY, align=AR, fmt=fmt or "General")


# ---------------------------------------------------------------- metas
def montar_metas(ws):
    ws.sheet_properties.tabColor = DARK
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 32, "B": 14, "C": 14, "D": 14, "E": 13, "F": 14})
    ws.row_dimensions[1].height = 34
    merged(ws, "A1:F1", "Metas Financeiras · Notinha", fill=F_TEAL, font=FT_TITLE)
    merged(ws, "A3:F3", "METAS (preencha Meta, Valor alvo, Já guardado e Prazo)")
    for col, h in zip("ABCDEF", ["Meta", "Valor alvo", "Já guardado", "Falta",
                                 "% concluído", "Prazo"]):
        ws[f"{col}4"] = h
    paint(ws, "A4:F4", fill=F_TEAL, font=FT_HEAD, align=AC)
    for r in range(5, 15):
        ws[f"D{r}"] = f"=B{r}-C{r}"
        ws[f"E{r}"] = f"=IFERROR(C{r}/B{r},0)"
    card_rows(ws, "A5:F14", {"B": FMT_BRL, "C": FMT_BRL, "D": FMT_BRL,
                             "E": FMT_PCT, "F": FMT_DATE})
    ws["A5"] = "Reserva de emergência (exemplo)"
    ws["B5"] = 5000.00
    ws["C5"] = 1500.00
    ws["F5"] = datetime(datetime.now().year, 12, 31)


# ---------------------------------------------------------------- investimento
def montar_investimento(ws):
    ws.sheet_properties.tabColor = DARK
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 16, "B": 14, "C": 14, "D": 14})
    ws.row_dimensions[1].height = 34
    merged(ws, "A1:D1", "Investimento · Notinha", fill=F_TEAL, font=FT_TITLE)
    merged(ws, "A3:D3", "INVESTIMENTOS POR MÊS (puxa das abas mensais)")
    for col, h in zip("ABCD", ["Mês", "Reserva", "Renda fixa", "Total"]):
        ws[f"{col}4"] = h
    paint(ws, "A4:D4", fill=F_TEAL, font=FT_HEAD, align=AC)
    # linhas de investimento na aba mensal
    n_saidas = len(SAIDAS_LINHAS)
    lin_tot_saidas = 31 + n_saidas
    inv_t = lin_tot_saidas + 2
    lin_reserva, lin_renda, lin_total = inv_t + 1, inv_t + 2, inv_t + 3
    for i, mes in enumerate(MESES):
        r = 5 + i
        ws[f"A{r}"] = mes
        ws[f"B{r}"] = f"={mes}!$I${lin_reserva}"
        ws[f"C{r}"] = f"={mes}!$I${lin_renda}"
        ws[f"D{r}"] = f"={mes}!$I${lin_total}"
    card_rows(ws, "A5:D16", {"B": FMT_BRL, "C": FMT_BRL, "D": FMT_BRL})
    ws["A17"] = "TOTAL DO ANO"
    for col in "BCD":
        ws[f"{col}17"] = f"=SUM({col}5:{col}16)"
    paint(ws, "A17:D17", fill=F_DARK, font=FT_HEAD, align=AR, fmt=FMT_BRL)
    ws["A17"].alignment = AL
    ws["A17"].number_format = "General"

    bar = BarChart()
    bar.type = "col"
    bar.title = "Total investido por mês"
    bar.add_data(Reference(ws, min_col=4, min_row=4, max_row=16), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=5, max_row=16))
    bar.legend = None
    bar.height, bar.width = 7.5, 13
    ws.add_chart(bar, "F3")


# ---------------------------------------------------------------- como usar
def montar_como_usar(ws):
    ws.sheet_properties.tabColor = DARK
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 4, "B": 100})
    ws.row_dimensions[1].height = 34
    merged(ws, "A1:B1", "Como usar · Notinha", fill=F_TEAL, font=FT_TITLE)
    linhas = [
        ("", None),
        ("O QUE A NOTINHA PREENCHE", F_TEAL),
        ("• O bloco \"Gastos do Mês\" (colunas A–E) de cada aba mensal: cada nota "
         "fiscal que você manda no WhatsApp vira uma linha, com Forma e Categoria "
         "já coloridas.", None),
        ("", None),
        ("O QUE VOCÊ PREENCHE", F_TEAL),
        ("• Fixos e Cartão de Crédito (colunas G–I).", None),
        ("• Entradas: Salário e Bônus.", None),
        ("• Saídas: só as linhas Reserva e Metas (o quanto você guardou). Débito, "
         "Crédito, Pix, Dinheiro e Voucher são somados automaticamente.", None),
        ("• Investimentos: Reserva e Renda fixa.", None),
        ("• Na tabela de categorias, \"Valor esperado\" é a sua meta de gasto por "
         "categoria.", None),
        ("• \"Contas a pagar\": suas faturas e boletos, com data de vencimento e o "
         "checkbox Pago? — é um lembrete do que ainda falta pagar.", None),
        ("", None),
        ("SAÍDAS × CONTAS A PAGAR (a diferença)", F_TEAL),
        ("• SAÍDAS por forma de pagamento = o que JÁ saiu no mês, somado por forma "
         "(Débito, Crédito, Pix...). É o que fecha o seu Saldo.", None),
        ("• CONTAS A PAGAR = o que ainda VAI sair (faturas/boletos a vencer). É um "
         "checklist, não entra no Saldo.", None),
        ("", None),
        ("COMO FUNCIONA", F_TEAL),
        ("• Cada aba mensal se atualiza sozinha: totais, saldo, categorias, "
         "insights e gráfico de pizza.", None),
        ("• Visão Anual, Metas e Investimento puxam os números das abas mensais.", None),
        ("• Forma e Categoria têm listas suspensas e viram \"pills\" coloridos "
         "automaticamente (as 12 categorias que a Notinha usa).", None),
        ("• Na Visão Anual, os blocos \"Top produtos\" e \"Maiores variações de "
         "preço\" são o diferencial item-a-item da Notinha.", None),
        ("", None),
        ("LEGENDA DE CORES", F_TEAL),
        ("Cabeçalho de bloco (teal Notinha)", F_TEAL),
        ("Faixa de destaque: totais e saldo", F_DARK),
        ("Célula de preenchimento", F_CARD),
    ]
    r = 3
    for texto, fill in linhas:
        cell = ws[f"B{r}"]
        cell.value = texto
        if fill is F_TEAL:
            cell.fill, cell.font = F_TEAL, FT_HEAD
        elif fill is F_DARK:
            cell.fill, cell.font = F_DARK, FT_HEAD
        elif fill is F_CARD:
            cell.fill, cell.font = F_CARD, FT_BODY
        else:
            cell.font = FT_BODY
        cell.alignment = AL
        r += 1


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("saida", nargs="?", default=None)
    ap.add_argument("--dados", default=None, help="JSON com notas reais (opcional)")
    ap.add_argument("--ano", type=int, default=None)
    args = ap.parse_args()

    dados = None
    if args.dados:
        dados = json.loads(Path(args.dados).read_text(encoding="utf-8"))
        if args.ano:
            dados["ano"] = args.ano

    out = Path(args.saida) if args.saida else (
        Path(__file__).resolve().parent.parent / "Notinha_Planilha_Mestre.xlsx")

    wb = Workbook()
    montar_como_usar(wb.active)
    wb.active.title = "Como usar"

    for i, mes in enumerate(MESES):
        montar_mes(wb.create_sheet(mes), i, dados)

    montar_visao_anual(wb.create_sheet("Visão Anual"), dados)
    montar_metas(wb.create_sheet("Metas Financeiras"))
    montar_investimento(wb.create_sheet("Investimento"))

    wb.save(out)
    print(f"Arquivo salvo: {out}")

    recalc = next((p for p in (
        Path("/mnt/skills/public/xlsx/scripts/recalc.py"),
        Path("/root/.claude/skills/xlsx/scripts/recalc.py"),
    ) if p.exists()), None)
    if recalc:
        r = subprocess.run([sys.executable, str(recalc), str(out), "180"],
                           capture_output=True, text=True)
        print(r.stdout.strip() or r.stderr.strip())
    else:
        print("AVISO: recalc.py não encontrado.")

    print("Abas:", load_workbook(out).sheetnames)


if __name__ == "__main__":
    main()
